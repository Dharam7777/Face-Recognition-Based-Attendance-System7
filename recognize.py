import cv2 as cv
from openpyxl import Workbook
import tensorflow as tf
import numpy as np
from datetime import datetime
import os
dir_path = os.path.dirname(os.path.realpath(__file__))
students=os.listdir('students')
num_classes=len(students)
def prepare(filepath):
		  IMG_SIZE=32
		  img_array=cv.imread(filepath,1)
		  e=cv.resize(img_array,(IMG_SIZE,IMG_SIZE))
		  img=e.copy()
		  img=img.astype('float')/255.0
		  img=np.expand_dims(img,axis=0)
		  return img.reshape(-1,IMG_SIZE,IMG_SIZE,3)
student=os.listdir('students')
CATEGORIES=list()
for label in student:
        CATEGORIES.append(label)
#CATEGORIES=["Dharam","Jainish","Parth"]
fp=dir_path+"\output"     
book=Workbook()
sheet=book.active
now= datetime.now()
today=now.day
month=now.month    
os.chdir(dir_path+"/model")
model=tf.keras.models.load_model("cnn"+str(num_classes)+".model")
os.chdir(fp)
os.chdir("..")
data=[]
students=os.listdir('output')
sheet.append(["Name","Day","Month","Time","Attendance"])
for label in students :
         for  i in os.listdir('output/'+label):		
             prediction=model.predict([prepare('output'+'/'+label+'/'+i)])
             result=CATEGORIES[np.argmax(prediction)]
             temp=list()
             temp=(result,today,month,now,'Present')
             data.append(temp)
             print(result)     
for row in data:
         max_row=sheet.max_row
         k=0
         for i in range (1,max_row+1):
             cell_obj=sheet.cell(row=i,column=1)
             cell_obj1=sheet.cell(row=i,column=2)
             if(cell_obj1.value==row[3] and cell_obj.value==row[0]):
                k=1     
         if(k==0):
          sheet.append(row)
os.chdir(dir_path+"/attendance sheet")   
book.save(str(month)+'.xlsx')  